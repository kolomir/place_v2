import openpyxl, os, dodatki
from datetime import date, datetime

file_path = "C:\\Users\\plmko\\Downloads\\Zestawienie nieobecności w roboczych.xlsx"

print(file_path)

teraz = datetime.today()
data_miesiac = str(dodatki.data_miesiac_dzis())

wb = openpyxl.load_workbook(os.path.join(file_path))
sheet = wb.active
used_columns = sheet.max_column

def convert_if_float(val):
    if val is None:
        return val
    # Jeśli wartość jest już typu float, zaokrąglamy i konwertujemy
    if isinstance(val, float):
        return int(round(val))
    # Jeśli wartość jest typu string, spróbujemy ją przekonwertować
    if isinstance(val, str):
        # Zamieniamy przecinek na kropkę, aby umożliwić konwersję
        temp = val.replace(',', '.')
        try:
            num = float(temp)
            return int(round(num))
        except ValueError:
            # Jeśli nie udało się przekonwertować, pozostawiamy oryginalny string
            return val
    # Inne typy pozostawiamy bez zmian
    return val

lista_wpisow = []

for row in sheet.iter_rows(min_row=13, min_col=0, max_col=30, values_only=True):
    # sprawdzamy czy wiersz nie jest pusty (zakładając że pusty wiersz ma wszystkie kolumny o wartosci None i kończy zestawienie)
    if any(cell is not None for cell in row):
        if row[4] == None:
            break
        else:
            tekst = row[4].split('|')
            nazwisko = tekst[0].strip()
            nr_akt = tekst[1].strip()

            wpis = [nazwisko, nr_akt, row[6], row[8], row[9], row[10], row[11], row[12], row[17], row[18], row[19], row[20], row[21], row[22], row[23], row[24], row[26], row[27], row[29], data_miesiac, teraz]
            wpis = [convert_if_float(x) for x in wpis]
            lista_wpisow.append(wpis)
            # print(nazwisko,nr_akt,row[6],row[8],row[9],row[10],row[11],row[12],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[26],row[27],row[29],data_miesiac,teraz)
    else:
        break

lista_wpisow_notNone = [tuple(0 if x is None else x for x in wiersz) for wiersz in lista_wpisow]

i=13
for row in lista_wpisow_notNone:
    print(i,' ',row)
    i+=1
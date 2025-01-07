import mysql.connector
from openpyxl import load_workbook
import math, os

# Funkcja do obcięcia liczb dziesiętnych
def truncate_float(value):
    try:
        return math.floor(float(value))
    except ValueError:
        return value  # Jeśli to nie jest liczba, zwraca wartość oryginalną

# Wczytanie danych z pliku Excel
def load_data_from_excel(file_path):
    workbook = load_workbook(os.path.join(file_path))
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data

# Zapis do bazy danych MySQL
def save_to_mysql(data):
    connection = mysql.connector.connect(
        host="localhost",  # Zmień na swój host
        user="root",       # Zmień na swoją nazwę użytkownika
        password="",  # Zmień na swoje hasło
        database="place_v2"  # Zmień na swoją bazę danych
    )
    cursor = connection.cursor()

    # Przygotowanie zapytania SQL
    for row in data[1:]:  # Pomijamy nagłówki
        truncated_row = [truncate_float(value) if isinstance(value, (int, float)) else value for value in row]
        placeholders = ", ".join(["%s"] * len(truncated_row))
        print(truncated_row)
        #sql = f"INSERT INTO {table_name} VALUES ({placeholders})"
        #cursor.execute(sql, truncated_row)

    connection.commit()
    cursor.close()
    connection.close()

# Ścieżka do pliku Excel
file_path = "D:\\ProjektyPython\\place_v2\\bledy\\styczeń\\testy - nie uzywac.xlsx"

# Wczytanie danych
data = load_data_from_excel(file_path)

# Zapis danych do MySQL
save_to_mysql(data)

print('---- KONIEC ----')

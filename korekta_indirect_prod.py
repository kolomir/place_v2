from PyQt5.QtWidgets import QWidget, QTableWidgetItem,QHeaderView,QApplication, QFileDialog, QMessageBox, QMenu, QCheckBox, QWidgetAction, QScrollArea, QPushButton, QFrame, QVBoxLayout
from PyQt5.QtCore import Qt, QPoint

from _korekta_indirect_prod_ui import Ui_Form
import db, dodatki
import openpyxl
from datetime import date, datetime
import os

from korekta_indirect_prod_dodaj import MainWindow_korekta_indirect_prod_dodaj
from linieFormEdytuj import MainWindow_linieEdytuj

# klasa która pozwala na poprawne sortowanie danych w kolumnach numerycznych.
# Normalnie dane układają się w sposób 1,10,11,2,23,245
# Po zastosowaniu klasy dane posortują się w sposób 1,2,10,11,23,245
class NumericTableWidgetItem(QTableWidgetItem):
    def __lt__(self, other):
        # Sprawdzamy, czy drugi element też jest instancją QTableWidgetItem
        if isinstance(other, QTableWidgetItem):
            try:
                # Porównujemy jako liczby
                return float(self.text()) < float(other.text())
            except ValueError:
                # W przypadku błędu porównujemy jako tekst
                return self.text() < other.text()
        return super().__lt__(other)

class MainWindow_korekta_indirect_prod(QWidget):
    def __init__(self, dzial):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        self.load_data_from_database()
        self.ui.btn_zapisz.clicked.connect(self.otworz_okno_korekta_indirect_prod_dodaj)
        self.ui.btn_przegladaj.clicked.connect(self.open_file_dialog)
        self.ui.btn_importuj.clicked.connect(self.czytaj_dane)
        self.ui.btn_szablon.clicked.connect(self.szablon)
        self.ui.tab_dane.itemChanged.connect(self.on_item_changed)

        self.ui.tab_dane.horizontalHeader().setSectionsClickable(True)
        self.ui.tab_dane.horizontalHeader().sectionClicked.connect(self.show_filter_menu)
        self.active_filters = {}  # Przechowywanie aktywnych filtrów dla każdej kolumny
        print('dzial 2',dzial)
        if dzial == 3:
            self.dzial_nazwa = "prod_cz"
        if dzial == 4:
            self.dzial_nazwa = "prod_bs"
        if dzial == 5:
            self.dzial_nazwa = "mag"

    def open_file_dialog(self):
        # Otwieranie dialogu wyboru pliku
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, "Wybierz plik Excel", "","Pliki tekstowe (*.xlsx);;Wszystkie pliki (*)", options=options)
        if file_path:
            self.ui.ed_sciezka_dane.setText(file_path)  # Ustawienie ścieżki w polu tekstowym

    def load_from_path(self):
        # Wczytanie pliku z ręcznie wpisanej ścieżki
        file_path = self.ui.ed_sciezka_dane.text()
        if file_path:
            self.load_file(file_path)

    def czytaj_dane(self):
        self.sprawdz_wpisy()
        file_path = self.ui.ed_sciezka_dane.text()
        wb = openpyxl.load_workbook(os.path.join(file_path))
        sheet = wb.active
        teraz = datetime.today()
        data_miesiac = str(dodatki.data_miesiac_dzis())
        print(data_miesiac)

        query = '''
                        select 
                            d.Nr_akt 
                            ,d.Nazwisko_i_imie 
                        from 
                            direct d 
                        where 
                            miesiac = '{0}'
                        '''.format(data_miesiac)
        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        results = db.read_query(connection, query)
        connection.close()

        lista_wpisow = []

        # czytamy wszystkie kolumny i wiersze ze wskazanego pliku
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
            # sprawdzamy czy wiersz nie jest pusty (zakładając że pusty wiersz ma wszystkie kolumny o wartosci None i kończy zestawienie)
            if any(cell is not None for cell in row):
                for dane in results:
                    if int(dane[0]) == row[0]:
                        pracownik = dane[1]
                lista_wpisow.append([row[0], pracownik, row[1], row[2], data_miesiac, teraz, self.dzial_nazwa])
            else:
                break

        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)

        for row in lista_wpisow:
            insert_data = "INSERT INTO korekta_indirect VALUES (NULL,'%s','%s','%s','%s','%s','%s','%s');" % (row[0], row[1], row[2], row[3], row[4], row[5], row[6])
            db.execute_query(connection, insert_data)

        self.ui.tab_dane.blockSignals(True)
        self.load_data_from_database()
        self.ui.tab_dane.blockSignals(False)

    def load_data_from_database(self):
        """Funkcja do załadowania danych z bazy do QTableWidget."""
        try:
            miestac_roboczy = dodatki.data_miesiac_dzis()
            select_data = "select ki.id, ki.nr_akt, ki.`nazwisko_i_imie`, ki.czas, ki.opis, ki.miesiac, ki.dzial from korekta_indirect ki where miesiac = '{0}';".format(miestac_roboczy)
            #select_data = "select * from kpi_mag"
            connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
            results = db.read_query(connection, select_data)

            self.ui.tab_dane.setSortingEnabled(True)

            self.ui.tab_dane.setColumnCount(6)  # Zmień na liczbę kolumn w twojej tabeli
            self.ui.tab_dane.setRowCount(0)  # Ustawienie liczby wierszy na 0
            self.ui.tab_dane.setHorizontalHeaderLabels([
                'Nr akt',
                'Nazwisko i imię',
                'Korekta',
                'Opis',
                'Data dodania',
                'Dział'
            ])

            # Ustawianie liczby wierszy na podstawie danych z bazy
            self.ui.tab_dane.setRowCount(len(results))

            # Wypełnianie tabeli danymi
            for row_idx, row_data in enumerate(results):
                # Przechowujemy id każdego wiersza
                for col_idx, value in enumerate(row_data[1:]):  # Pomijamy id
                    item = NumericTableWidgetItem(str(value))              # Użycie klasy soryującej dane numeryczne

                    item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)

                    self.ui.tab_dane.setColumnWidth(0, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(1, 200)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(2, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(3, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(4, 150)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(5, 100)  # Stała szerokość: 150 pikseli

                    if col_idx == 0 or col_idx == 1 or col_idx == 5:  # Zablokowanie edycji dla kolumny "nazwa"
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Usuwamy flagę edytowalności
                    else:
                        item.setFlags(item.flags() | Qt.ItemIsEditable)  # Ustawienie komórek jako edytowalne

                    self.ui.tab_dane.setItem(row_idx, col_idx, item)

            # Przechowywanie id wierszy
            self.row_ids = [row_data[0] for row_data in results]
            #print(row_data[0] for row_data in results)

        except db.Error as e:
            print(f"Błąd przy pobieraniu danych z bazy danych: {e}")

    def update_database(self, record_id, col, new_value):
        """Funkcja do aktualizacji konkretnej komórki w bazie danych."""
        try:
            # Mapowanie indeksu kolumny na nazwę kolumny w bazie
            column_names = ["nr_akt", "nazwisko_i_imie", "czas", "opis"]
            column_name = column_names[col]
            if col == 2:
                new_value = new_value.replace(",", ".")

            # Aktualizacja w bazie danych
            sql_query = f"UPDATE korekta_indirect SET {column_name} = %s WHERE id = %s"
            connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
            db.execute_query_virable(connection,sql_query,(new_value, record_id))

            #self.sprawdz_dzial()
            #self.cursor.execute(sql_query, (new_value, record_id))
            #self.db_connection.commit()
            #print(f"Zaktualizowano rekord o id {record_id}, {column_name} = {new_value}")

        except db.Error as e:
            print(f"Błąd zapisu do bazy danych: {e}")

    #def sprawdz_dzial(self, item):
    #    row = item.row()
    #    record_id = self.row_ids[row]

    #    select_data = "select ki.id, ki.dzial from korekta_indirect ki where id = '{0}';".format(record_id)
    #    connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
    #    results = db.read_query(connection, select_data)
    #    print('Sprawdzenie: ',results[0],'; ',results[1])


    def on_item_changed(self, item):
        """Funkcja wywoływana przy każdej zmianie komórki."""
        row = item.row()
        col = item.column()
        new_value = item.text()

        # Pobranie id rekordu dla zmienionego wiersza
        record_id = self.row_ids[row]

        select_data = "select ki.id, ki.dzial from korekta_indirect ki where id = {0};".format(record_id)
        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        results = db.read_query(connection, select_data)
        #print('Sprawdzenie: ', results)
        #print('Sprawdzenie: ', results[0][0], '; ', results[0][1], '; ', self.dzial_nazwa)

        if results[0][1] == self.dzial_nazwa:
            # Zapis zmienionych danych do bazy
            self.update_database(record_id, col, new_value)
        else:
            QMessageBox.critical(self, 'Error', 'Prawdopodobnie to nie jest Twój pracownik! <br />Wybierz inne nazwisko do korekty czasu.')
            return

    def otworz_okno_korekta_indirect_prod_dodaj(self):
        dzial = self.dzial_nazwa
        self.okno_korekta_indirect_prod_dodaj = MainWindow_korekta_indirect_prod_dodaj(dzial)
        self.okno_korekta_indirect_prod_dodaj.show()

    def szablon(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        headers = ["nr prac.","czas","opis"]
        ws.append(headers)

        # Ustawianie szerokości kolumn
        ws.column_dimensions['A'].width = 5  # Kolumna 'lp'
        ws.column_dimensions['B'].width = 10  # Kolumna 'nr pracownika'
        ws.column_dimensions['C'].width = 65  # Kolumna 'IMIĘ i NAZWISKO'

        domyslna_nazwa = 'korekta_direct.xlsx'
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            None,
            "Zapisz plik",
            domyslna_nazwa,  # Domyślna nazwa pliku
            "Pliki Excel (*.xlsx);;Wszystkie pliki (*)",
            options=options
        )

        # Sprawdzanie, czy użytkownik wybrał plik
        if file_path:
            wb.save(file_path)
            print(f"Plik zapisano: {file_path}")
        else:
            print("Zapis pliku anulowany")

    def sprawdz_wpisy(self):
        miestac_roboczy = dodatki.data_miesiac_dzis()
        select_data = "SELECT * FROM `korekta_indirect` WHERE miesiac = '%s' and dzial = '%s';" % (miestac_roboczy, self.dzial_nazwa)  # (miestac_roboczy)
        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        results = db.read_query(connection, select_data)
        connection.close()

        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        if results:
            for x in results:
                delete_data = "delete from korekta_indirect where id = '%s' and miesiac = '%s' and dzial = '%s';" % (x[0],miestac_roboczy, self.dzial_nazwa)
                print('Do skasowania:',delete_data)
                db.execute_query(connection, delete_data)
        else:
            print('--Brak wpisów jeszcze--')
        connection.close()

    def show_filter_menu(self, col):
        values = set(
            self.ui.tab_dane.item(row, col).text()
            for row in range(self.ui.tab_dane.rowCount())
            if self.ui.tab_dane.item(row, col)
        )

        menu = QMenu(self)

        # Tworzenie obszaru przewijania
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_content = QFrame()
        layout = QVBoxLayout(scroll_content)

        checkboxes = {}
        max_width = 200  # Minimalna szerokość
        for value in sorted(values):
            checkbox = QCheckBox(value)
            checkbox.setChecked(
                col not in self.active_filters or value in self.active_filters[col]
            )
            layout.addWidget(checkbox)
            checkboxes[checkbox] = value

            # Dostosowanie szerokości na podstawie tekstu
            width = checkbox.fontMetrics().boundingRect(value).width() + 30
            max_width = min(max(max_width, width), 800)  # Maksymalna szerokość 800

        # Dodanie przycisków "Zaznacz wszystko" i "Wyczyść wszystko"
        button_select_all = QPushButton("Zaznacz wszystko")
        button_clear_all = QPushButton("Wyczyść wszystko")
        layout.addWidget(button_select_all)
        layout.addWidget(button_clear_all)

        button_select_all.clicked.connect(lambda: self.set_all_checkboxes(checkboxes, True))
        button_clear_all.clicked.connect(lambda: self.set_all_checkboxes(checkboxes, False))

        scroll_area.setWidget(scroll_content)
        scroll_area.setFixedHeight(200)  # Ograniczenie widocznych elementów do około 10 pozycji
        scroll_area.setMinimumWidth(max_width)
        scroll_area.setMaximumWidth(800)

        # Dodanie przewijanego obszaru do menu
        scroll_action = QWidgetAction(self)
        scroll_action.setDefaultWidget(scroll_area)
        menu.addAction(scroll_action)

        apply_action = menu.addAction("Zastosuj filtr")
        menu.addSeparator()
        clear_action = menu.addAction("Wyczyść filtr")

        header_pos = self.ui.tab_dane.mapToGlobal(self.ui.tab_dane.horizontalHeader().pos())
        section_pos = self.ui.tab_dane.horizontalHeader().sectionPosition(col)
        menu_pos = header_pos + QPoint(section_pos, self.ui.tab_dane.horizontalHeader().height())
        selected_action = menu.exec(menu_pos)

        if selected_action == apply_action:
            selected_values = [value for checkbox, value in checkboxes.items() if checkbox.isChecked()]
            self.active_filters[col] = selected_values
            self.apply_filters()
        elif selected_action == clear_action:
            if col in self.active_filters:
                del self.active_filters[col]
            self.apply_filters()

    def set_all_checkboxes(self, checkboxes, state):
        for checkbox in checkboxes.keys():
            checkbox.setChecked(state)

    def apply_filters(self):
        for row in range(self.ui.tab_dane.rowCount()):
            show_row = True
            for col, filter_values in self.active_filters.items():
                item = self.ui.tab_dane.item(row, col)
                if item and item.text() not in filter_values:
                    show_row = False
                    break
            self.ui.tab_dane.setRowHidden(row, not show_row)
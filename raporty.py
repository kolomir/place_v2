from PyQt5.QtWidgets import QWidget, QFileDialog
from PyQt5 import QtGui

import openpyxl
from openpyxl.styles import Alignment
from datetime import date, datetime, timedelta

from _raporty_ui import Ui_Form
import db, dodatki

class MainWindow_raporty(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        self.sprzwdzenie_raportow()

        self.ui.btn_eksport_enova.clicked.connect(self.raport_eksport_enova)
        self.ui.btn_zestawienie_pracownicy.clicked.connect(self.raport_zestawienie)

    def sprzwdzenie_raportow(self):
        data_miesiac = str(dodatki.data_miesiac_dzis())
        select_data = "SELECT * FROM zestawienia_prod zp WHERE zp.miesiac = '{0}'".format(data_miesiac)
        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        result = db.read_query(connection, select_data)
        connection.close()

        select_data_eksport = "SELECT * FROM eksport_danych ed WHERE ed.miesiac = '{0}'".format(
            data_miesiac)
        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        result_eksport = db.read_query(connection, select_data_eksport)
        connection.close()

        licz_cz = licz_bs = 0
        for dane in result_eksport:
            if dane[6] == 'prod':
                licz_cz += 1
            if dane[6] == 'mag':
                licz_bs += 1

        if not result:
            self.ui.lab_dot_zestawienie_pracownicy.setPixmap(QtGui.QPixmap(":/icon/img/svg_icons/dot_red.svg"))
        else:
            self.ui.lab_dot_zestawienie_pracownicy.setPixmap(QtGui.QPixmap(":/icon/img/svg_icons/dot_green.svg"))

        if licz_cz == 0:
            self.ui.lab_dot_eksport_enova_cz.setPixmap(QtGui.QPixmap(":/icon/img/svg_icons/dot_red.svg"))
        else:
            self.ui.lab_dot_eksport_enova_cz.setPixmap(QtGui.QPixmap(":/icon/img/svg_icons/dot_green.svg"))

        if licz_bs == 0:
            self.ui.lab_dot_eksport_enova_bs.setPixmap(QtGui.QPixmap(":/icon/img/svg_icons/dot_red.svg"))
        else:
            self.ui.lab_dot_eksport_enova_bs.setPixmap(QtGui.QPixmap(":/icon/img/svg_icons/dot_green.svg"))

    def raport_eksport_enova(self):
        data_miesiac = str(dodatki.data_miesiac_dzis())
        select_data = "SELECT * FROM eksport_danych zp WHERE zp.miesiac = '{0}'".format(data_miesiac)
        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        result = db.read_query(connection, select_data)
        connection.close()

        dzisiaj = datetime.today()
        pierwszy_dzien_biezacego_miesiaca = dzisiaj.replace(day=1)
        ostatni_dzien_biezacego_miesiaca = (pierwszy_dzien_biezacego_miesiaca + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        pierwszy_dzien_poprzedniego_miesiaca = (pierwszy_dzien_biezacego_miesiaca - timedelta(days=1)).replace(day=1)
        ostatni_dzien_poprzedniego_miesiaca = pierwszy_dzien_biezacego_miesiaca - timedelta(days=1)

        wb = openpyxl.Workbook()
        ws = wb.active

        pierwszy = pierwszy_dzien_poprzedniego_miesiaca.strftime("%d.%m.%Y")
        ostatni = ostatni_dzien_poprzedniego_miesiaca.strftime("%d.%m.%Y")
        print(pierwszy)
        print(ostatni)
        print(pierwszy_dzien_poprzedniego_miesiaca.strftime("%d.%m.%Y"))
        print(ostatni_dzien_poprzedniego_miesiaca.strftime("%d.%m.%Y"))

        headers = ["Pracownik:Class", "Pracownik:Kod", "Last.Okres.Od", "Last.Okres.Do", "Last.Element:Nazwa", "Last.Podstawa"]

        # Ustawianie szerokości kolumn
        ws.column_dimensions['A'].width = 16  # Kolumna 'Pracownik:Class'
        ws.column_dimensions['B'].width = 15  # Kolumna 'Pracownik:Kod'
        ws.column_dimensions['C'].width = 14  # Kolumna 'Last.Okres.Od'
        ws.column_dimensions['D'].width = 14  # Kolumna 'Last.Okres.Do'
        ws.column_dimensions['E'].width = 20  # Kolumna 'Last.Element:Nazwa'
        ws.column_dimensions['F'].width = 14  # Kolumna 'Last.Podstawa'

        lista_eksport = []
        lista_eksport.append(headers)
        for dane in result:
            #kwota = float(dane[4]).replace('.', ',')
            #kwota_liczbowa = float(kwota.replace(',', '.'))

            #print(['PracownikFirmy', dane[2], pierwszy, ostatni,'Premia za Produkt.',dane[4]])
            lista_eksport.append(['PracownikFirmy', dane[2], pierwszy, ostatni,'Premia za Produkt.',dane[4]])


        for row in lista_eksport:
            new_row = list(row)
            ws.append(new_row)

        domyslna_nazwa = 'eksport_Enova.xlsx'
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            None,
            "Zapisz plik",
            domyslna_nazwa,  # Domyślna nazwa pliku
            "Pliki Excel (*.xlsx);;Wszystkie pliki (*)",
            options=options
        )

        # Sprawdzanie, czy użytkownik wybrał plik
        if file_path:
            wb.save(file_path)
            print(f"Plik zapisano: {file_path}")
        else:
            print("Zapis pliku anulowany")

        #output_file = "output.xlsx"
        #wb.save(output_file)

    def raport_zestawienie(self):
        data_miesiac = str(dodatki.data_miesiac_dzis())
        select_data = "SELECT * FROM zestawienia_prod zp WHERE zp.miesiac = '{0}'".format(data_miesiac)
        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        result = db.read_query(connection, select_data)
        connection.close()

        wb = openpyxl.Workbook()
        ws = wb.active

        dzisiaj = datetime.today()
        pierwszy_dzien_biezacego_miesiaca = dzisiaj.replace(day=1)
        pierwszy = pierwszy_dzien_biezacego_miesiaca.strftime("%d.%m.%Y")

        headers = ["Nazwisko i Imię", "Nr akt", "Linia", "Data", "Nazwa Direct", "Direct %", "Nazwa IW", "IW %", "Nazwa Wydajność", "Wydajność", "Nazwa Produktywność", "Produktywność", "Nazwa Błędy", "Błędy", "Nazwa Nieobecność", "Nieobecność"]

        # Ustawianie szerokości kolumn
        ws.column_dimensions['A'].width = 38  # Kolumna 'Nazwisko i Imię'
        ws.column_dimensions['B'].width = 7  # Kolumna 'Nr akt'
        ws.column_dimensions['C'].width = 6  # Kolumna 'Linia'
        ws.column_dimensions['D'].width = 11  # Kolumna 'Data'
        ws.column_dimensions['E'].width = 13  # Kolumna 'Nazwa Direct'
        ws.column_dimensions['F'].width = 9  # Kolumna 'Direct %'
        ws.column_dimensions['G'].width = 14  # Kolumna 'Nazwa IW'
        ws.column_dimensions['H'].width = 9  # Kolumna 'IW %'
        ws.column_dimensions['I'].width = 18  # Kolumna 'Nazwa Wydajność'
        ws.column_dimensions['J'].width = 11  # Kolumna 'Wydajność'
        ws.column_dimensions['K'].width = 22  # Kolumna 'Nazwa Produktywność'
        ws.column_dimensions['L'].width = 16  # Kolumna 'Produktywność'
        ws.column_dimensions['M'].width = 13  # Kolumna 'Nazwa Błędy'
        ws.column_dimensions['N'].width = 7  # Kolumna 'Błędy'
        ws.column_dimensions['O'].width = 20  # Kolumna 'Nazwa Nieobecność'
        ws.column_dimensions['P'].width = 13  # Kolumna 'Nieobecność'ws[f"{col}1"].alignment = Alignment(vertical='center')

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.alignment = Alignment(vertical='center')



        lista_pracownicy = []
        #lista_pracownicy.append(headers)
        for dane in result:
            direkt = str(float(dane[4]) * 100.00)
            indirect = str(float(dane[5]) * 100.00)
            lista_pracownicy.append([dane[2], dane[1], dane[3], pierwszy, "Direct Work", f"{direkt}%", "Indirect Work", f"{indirect}%", "Wydajność", f"{dane[6]}%", "Produktywność", f"{dane[7]}%", "Błędy", dane[9], "Nieobecność", dane[8]])

        wiersz = 1
        for row in lista_pracownicy:
            new_row = list(row)
            ws.append(new_row)
            ws.row_dimensions[wiersz].height = 30
            wiersz += 1

        for row2 in ws.iter_rows():
            for cell in row2:
                cell.alignment = Alignment(vertical='center')

        domyslna_nazwa = 'raportowanie_pracownicy.xlsx'
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            None,
            "Zapisz plik",
            domyslna_nazwa,  # Domyślna nazwa pliku
            "Pliki Excel (*.xlsx);;Wszystkie pliki (*)",
            options=options
        )

        # Sprawdzanie, czy użytkownik wybrał plik
        if file_path:
            wb.save(file_path)
            print(f"Plik zapisano: {file_path}")
        else:
            print("Zapis pliku anulowany")

        #output_file = "output_pracownicy.xlsx"
        #wb.save(output_file)
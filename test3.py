from openpyxl import load_workbook
import os, math

# Ścieżka do pliku Excel
file_path = "D:\\ProjektyPython\\place_v2\\bledy\\styczeń\\testy - nie uzywac.xlsx"

def truncate_float(value):
    try:
        return math.floor(float(value))
    except ValueError:
        return value  # Jeśli to nie jest liczba, zwraca wartość oryginalną

def load_data_from_excel(file_path):
    workbook = load_workbook(os.path.join(file_path))
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data

def save_to_mysql(data):

    lista_wpisow_notNone = [tuple(0 if x is None else x for x in wiersz) for wiersz in data[1:]]
    # Przygotowanie zapytania SQL
    for row in lista_wpisow_notNone:  # Pomijamy nagłówki

        truncated_row = [truncate_float(value) if isinstance(value, (int, float)) else value for value in row]
        placeholders = ", ".join(["%s"] * len(truncated_row))
        print(truncated_row[1],truncated_row[2])

        #sql = f"INSERT INTO {table_name} VALUES ({placeholders})"
        #cursor.execute(sql, truncated_row)

# Wczytanie danych
data = load_data_from_excel(file_path)

save_to_mysql(data)
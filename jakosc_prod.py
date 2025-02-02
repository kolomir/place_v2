from PyQt5.QtWidgets import QWidget, QMessageBox, QFileDialog, QTableWidgetItem,QHeaderView, QMenu, QCheckBox, QWidgetAction, QScrollArea, QPushButton, QFrame, QVBoxLayout
from PyQt5.QtCore import Qt, QPoint
import configparser
import openpyxl
#import sys
import os
from datetime import date, datetime
import calendar

from jakosc_prodDodaj import MainWindow_jakosc_prodDodaj
from _jakosc_prod_ui import Ui_Form
import db, dodatki

# klasa która pozwala na poprawne sortowanie danych w kolumnach numerycznych.
# Normalnie dane układają się w sposób 1,10,11,2,23,245
# Po zastosowaniu klasy dane posortują się w sposób 1,2,10,11,23,245
class NumericTableWidgetItem(QTableWidgetItem):
    def __lt__(self, other):
        # Sprawdzamy, czy drugi element też jest instancją QTableWidgetItem
        if isinstance(other, QTableWidgetItem):
            try:
                # Porównujemy jako liczby
                return float(self.text()) < float(other.text())
            except ValueError:
                # W przypadku błędu porównujemy jako tekst
                return self.text() < other.text()
        return super().__lt__(other)

class MainWindow_jakosc(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        # - wczytanie pliku INI --------------------------------------------
        self.config = configparser.ConfigParser()
        self.config.read('config.ini')
        # - załadowanie zmiennych z pliku INI ------------------------------
        self.folder_bledy = self.config['sciezki']['folder_bledy']
        self.plik = self.config['sciezki']['plik_jakosc']
        # ------------------------------------------------------------------
        # - domyślna ścieżka dla pliku -----------------
        # domyslny = f"{self.folder_bledy}"
        # self.ui.ed_sciezka_dane.setText(domyslny)
        # -----------------------------------------------

        self.ui.btn_przegladaj.clicked.connect(self.open_file_dialog)
        self.ui.btn_importuj.clicked.connect(self.czytaj_dane)
        self.ui.btn_dodaj.clicked.connect(self.otworz_okno_jakosc_prodDodaj)
        self.ui.btn_szablon.clicked.connect(self.szablon)

        self.load_data_from_database()
        self.ui.tab_dane.itemChanged.connect(self.on_item_changed)

    def open_file_dialog(self):
        # Otwieranie dialogu wyboru pliku
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, "Wybierz plik Excel", "", "Pliki tekstowe (*.xlsx);;Wszystkie pliki (*)", options=options)
        if file_path:
            self.ui.ed_sciezka_dane.setText(file_path)  # Ustawienie ścieżki w polu tekstowym

    def folder_istnieje(self):
        folder = self.ui.ed_sciezka_dane.text().strip()
        self.flaga_import = 1
        if not folder:
            QMessageBox.critical(self, 'Error', 'Nie wybrano lokalizacji pliku')
            self.ui.ed_sciezka_dane.setFocus()
            return False

        if not os.path.exists(folder):
            QMessageBox.critical(self, 'Error', 'Folder nie istnieje. Sprawdź lokalizację pliku')
            self.ui.ed_sciezka_dane.setFocus()
            return False
        return True

    def load_from_path(self):
        # Wczytanie pliku z ręcznie wpisanej ścieżki
        file_path = self.ui.ed_sciezka_dane.text()
        if file_path:
            self.load_file(file_path)

    def sprawdz_wpisy(self):
        miestac_roboczy = dodatki.data_miesiac_dzis()
        select_data = "SELECT * FROM `jakosc_prod` WHERE miesiac = '%s';" % (miestac_roboczy)  # (miestac_roboczy)
        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        results = db.read_query(connection, select_data)
        connection.close()

        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        if results:
            for x in results:
                delete_data = "delete from jakosc_prod where id = '%s' and miesiac = '%s';" % (x[0],miestac_roboczy)
                print('Do skasowania:',delete_data)
                db.execute_query(connection, delete_data)
        else:
            print('--Brak wpisów jeszcze--')
        connection.close()

    def czytaj_dane(self):
        if not self.folder_istnieje():
            return
        self.sprawdz_wpisy()
        file_path = self.ui.ed_sciezka_dane.text()
        wb = openpyxl.load_workbook(os.path.join(file_path))
        sheet = wb.active
        teraz = datetime.today()
        data_miesiac = str(dodatki.data_miesiac_dzis())
        # print(data_miesiac)

        lista_wpisow = []

        i = 1
        # czytamy wszystkie kolumny i wiersze ze wskazanego pliku
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
            # sprawdzamy czy wiersz nie jest pusty (zakładając że pusty wiersz ma wszystkie kolumny o wartosci None i kończy zestawienie)
            if any(cell is not None for cell in row):
                grupa = row[0].strip()
                grupa_robocza = str(row[1]).strip()
                ppm = str(row[2]).strip()
                reklamacje = str(row[3]).strip()
                lista_wpisow.append([grupa, grupa_robocza, ppm, reklamacje, data_miesiac, teraz])
            else:
                break

        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)

        for row in lista_wpisow:
            insert_data = "INSERT INTO jakosc_prod VALUES (NULL,'%s','%s','%s','%s','%s','%s');" % (
            row[0], row[1], row[2], row[3], row[4], row[5])
            db.execute_query(connection, insert_data)

        self.ui.tab_dane.blockSignals(True)
        self.load_data_from_database()
        self.ui.tab_dane.blockSignals(False)

    def load_data_from_database(self):
        """Funkcja do załadowania danych z bazy do QTableWidget."""
        try:
            miestac_roboczy = dodatki.data_miesiac_dzis()
            select_data = "select * from jakosc_prod where miesiac = '%s';" % (miestac_roboczy)
            #select_data = "select * from kpi_mag"
            connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
            results = db.read_query(connection, select_data)

            self.ui.tab_dane.setSortingEnabled(True)

            self.ui.tab_dane.setColumnCount(6)  # Zmień na liczbę kolumn w twojej tabeli
            self.ui.tab_dane.setRowCount(0)  # Ustawienie liczby wierszy na 0
            self.ui.tab_dane.setHorizontalHeaderLabels([
                'Grupa',
                'Grupa robocza',
                'PPM',
                'Reklamacje',
                'Miesiac',
                'Data dodania'
            ])

            # Ustawianie liczby wierszy na podstawie danych z bazy
            self.ui.tab_dane.setRowCount(len(results))

            # Wypełnianie tabeli danymi
            for row_idx, row_data in enumerate(results):
                # Przechowujemy id każdego wiersza
                for col_idx, value in enumerate(row_data[1:]):  # Pomijamy id
                    item = NumericTableWidgetItem(str(value))              # Użycie klasy soryującej dane numeryczne

                    item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)

                    self.ui.tab_dane.setColumnWidth(0, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(1, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(2, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(3, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(4, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(5, 150)  # Stała szerokość: 150 pikseli

                    if col_idx == 2 or col_idx == 3:  # Zablokowanie edycji dla kolumny "nazwa"
                        item.setFlags(item.flags() | Qt.ItemIsEditable)  # Ustawienie komórek jako edytowalne
                    else:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Usuwamy flagę edytowalności

                    self.ui.tab_dane.setItem(row_idx, col_idx, item)

            # Przechowywanie id wierszy
            self.row_ids = [row_data[0] for row_data in results]
            #print(row_data[0] for row_data in results)

        except db.Error as e:
            print(f"Błąd przy pobieraniu danych z bazy danych: {e}")

    def on_item_changed(self, item):
        """Funkcja wywoływana przy każdej zmianie komórki."""
        row = item.row()
        col = item.column()
        #print(f"DEBUG: Wybrane kolumny: {col}")
        new_value = item.text()

        # Pobranie id rekordu dla zmienionego wiersza
        record_id = self.row_ids[row]
        #print('record_id:',record_id)

        # Zapis zmienionych danych do bazy
        self.update_database(record_id, col, new_value)


    def update_database(self, record_id, col, new_value):
        """Funkcja do aktualizacji konkretnej komórki w bazie danych."""
        try:
            # Mapowanie indeksu kolumny na nazwę kolumny w bazie
            column_names = ["grupa", "grupa_robocza", "ppm", "reklamacje"]
            column_name = column_names[col]

            # Aktualizacja w bazie danych
            sql_query = f"UPDATE jakosc_prod SET {column_name} = %s WHERE id = %s"
            connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
            db.execute_query_virable(connection,sql_query,(new_value, record_id))
            #print(f"Zaktualizowano rekord o id {record_id}, {column_name} = {new_value}")

        except db.Error as e:
            print(f"Błąd zapisu do bazy danych: {e}")

    def szablon(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        headers = ["Grupa","Grupa robocza","PPM","Reklamacje"]
        ws.append(headers)

        line1 = ["lider","3011","",""]
        line2 = ["lider","3012","",""]
        line3 = ["lider","3013+3015","",""]
        line4 = ["lider","3014","",""]
        line5 = ["lider","3016+3019","",""]
        line6 = ["lider","3018","",""]
        line7 = ["lider","3031+3032","",""]
        line8 = ["lider","Montaż","",""]
        line9 = ["lider","Zakuwanie","",""]
        line10 = ["lider","Maszyny","",""]
        line11 = ["instruktor","Czaplinek","",""]
        line12 = ["instruktor","Borne Sulinowo","",""]

        ws.append(line1)
        ws.append(line2)
        ws.append(line3)
        ws.append(line4)
        ws.append(line5)
        ws.append(line6)
        ws.append(line7)
        ws.append(line8)
        ws.append(line9)
        ws.append(line10)
        ws.append(line11)
        ws.append(line12)


        # Ustawianie szerokości kolumn
        ws.column_dimensions['A'].width = 15  # Grupa
        ws.column_dimensions['B'].width = 15  # Grupa robocza
        ws.column_dimensions['C'].width = 10  # PPM
        ws.column_dimensions['D'].width = 7  # Reklamacje

        domyslna_nazwa = 'jakosc.xlsx'
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            None,
            "Zapisz plik",
            domyslna_nazwa,  # Domyślna nazwa pliku
            "Pliki Excel (*.xlsx);;Wszystkie pliki (*)",
            options=options
        )

        # Sprawdzanie, czy użytkownik wybrał plik
        if file_path:
            wb.save(file_path)
            print(f"Plik zapisano: {file_path}")
        else:
            print("Zapis pliku anulowany")

    def otworz_okno_jakosc_prodDodaj(self):
        self.okno_jakosc_prodDodaj = MainWindow_jakosc_prodDodaj()
        self.okno_jakosc_prodDodaj.show()
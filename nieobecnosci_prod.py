from PyQt5.QtWidgets import QWidget, QMessageBox, QFileDialog, QTableWidgetItem,QHeaderView, QMenu, QCheckBox, QWidgetAction, QScrollArea, QPushButton, QFrame, QVBoxLayout
from PyQt5.QtCore import Qt, QPoint
import configparser
import openpyxl
#import sys
import os, math
from datetime import date, datetime
import calendar

from plik_pomoc_nieobecnosci import MainWindow_pomoc_nieobecnosci
from _nieobecnosci_prod_ui import Ui_Form
import db, dodatki

# klasa która pozwala na poprawne sortowanie danych w kolumnach numerycznych.
# Normalnie dane układają się w sposób 1,10,11,2,23,245
# Po zastosowaniu klasy dane posortują się w sposób 1,2,10,11,23,245
class NumericTableWidgetItem(QTableWidgetItem):
    def __lt__(self, other):
        # Sprawdzamy, czy drugi element też jest instancją QTableWidgetItem
        if isinstance(other, QTableWidgetItem):
            try:
                # Porównujemy jako liczby
                return float(self.text()) < float(other.text())
            except ValueError:
                # W przypadku błędu porównujemy jako tekst
                return self.text() < other.text()
        return super().__lt__(other)

class MainWindow_nieobecnosci(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        self.load_data_from_database()
        self.ui.btn_szablon_p.clicked.connect(self.otworz_okno_plik_pomoc_nieobecnosci)
        self.ui.btn_szablon_o.clicked.connect(self.szablon_obco)
        self.ui.btn_przegladaj.clicked.connect(self.open_file_dialog)
        self.ui.btn_importuj.clicked.connect(self.czytaj_dane)
        self.ui.btn_importuj_obco.clicked.connect(self.czytaj_dane_obco)

        self.ui.tab_dane.horizontalHeader().setSectionsClickable(True)
        self.ui.tab_dane.horizontalHeader().sectionClicked.connect(self.show_filter_menu)
        self.active_filters = {}  # Przechowywanie aktywnych filtrów dla każdej kolumny

    def load_data_from_database(self):
        """Funkcja do załadowania danych z bazy do QTableWidget."""
        try:
            miestac_roboczy = dodatki.data_miesiac_dzis()
            select_data = "SELECT * FROM `nieobecnosci_prod` WHERE miesiac = '%s';" % (miestac_roboczy)
            # select_data = "select * from kpi_mag"
            connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
            results = db.read_query(connection, select_data)

            self.ui.tab_dane.setSortingEnabled(True)

            self.ui.tab_dane.setColumnCount(22)  # Zmień na liczbę kolumn w twojej tabeli
            self.ui.tab_dane.setRowCount(0)  # Ustawienie liczby wierszy na 0
            self.ui.tab_dane.setHorizontalHeaderLabels([
                'Nazwisko i imię',
                'Nr akt',
                'Stanowisko',
                'Urlop wypocz.',
                'Urlop bezpłatny',
                'Urlop szkoleniowy',
                'Urlop opieka (art. 188 kp) ',
                'Urlop okoliczność.',
                'Zwol. lek.',
                'Urlop macierz.',
                'Opieka ZUS',
                'Urlop wychow.',
                'Inne nieobecn.',
                'Usp.',
                'NN',
                'Rehab.',
                'Rodz.',
                'Krew',
                'Dni w pracy',
                'Razem',
                'Miesiac',
                'Data dodania'
            ])

            # Ustawianie liczby wierszy na podstawie danych z bazy
            self.ui.tab_dane.setRowCount(len(results))

            # Wypełnianie tabeli danymi
            for row_idx, row_data in enumerate(results):
                # Przechowujemy id każdego wiersza
                for col_idx, value in enumerate(row_data[1:]):  # Pomijamy id
                    item = NumericTableWidgetItem(str(value))              # Użycie klasy soryującej dane numeryczne

                    item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)

                    self.ui.tab_dane.setColumnWidth(0, 150)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(1, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(2, 200)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(3, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(4, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(5, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(6, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(7, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(8, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(9, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(10, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(11, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(12, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(13, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(14, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(15, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(16, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(17, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(18, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(19, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(20, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(21, 150)  # Stała szerokość: 150 pikseli

                    self.ui.tab_dane.setItem(row_idx, col_idx, item)

            # Przechowywanie id wierszy
            self.row_ids = [row_data[0] for row_data in results]
            #print(row_data[0] for row_data in results)

        except db.Error as e:
            print(f"Błąd przy pobieraniu danych z bazy danych: {e}")

    def otworz_okno_plik_pomoc_nieobecnosci(self):
        self.okno_plik_pomoc_nieobecnosci = MainWindow_pomoc_nieobecnosci()
        self.okno_plik_pomoc_nieobecnosci.show()

    def szablon_obco(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        headers = ["nr_akt","Nazwisko i imie","dni w pracy"]
        ws.append(headers)

        # Ustawianie szerokości kolumn
        ws.column_dimensions['A'].width = 8  # Kolumna 'nr pracownika'
        ws.column_dimensions['B'].width = 30  # Kolumna 'IMIĘ i NAZWISKO'
        ws.column_dimensions['C'].width = 7  # Kolumna 'Suma z błędów'

        domyslna_nazwa = 'nieobecnosci_obcokrajowcow.xlsx'
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            None,
            "Zapisz plik",
            domyslna_nazwa,  # Domyślna nazwa pliku
            "Pliki Excel (*.xlsx);;Wszystkie pliki (*)",
            options=options
        )

        # Sprawdzanie, czy użytkownik wybrał plik
        if file_path:
            wb.save(file_path)
            print(f"Plik zapisano: {file_path}")
        else:
            print("Zapis pliku anulowany")

    def open_file_dialog(self):
        # Otwieranie dialogu wyboru pliku
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, "Wybierz plik Excel", "", "Pliki tekstowe (*.xlsx);;Wszystkie pliki (*)", options=options)
        if file_path:
            self.ui.ed_sciezka_dane.setText(file_path)  # Ustawienie ścieżki w polu tekstowym

    def load_from_path(self):
        # Wczytanie pliku z ręcznie wpisanej ścieżki
        file_path = self.ui.ed_sciezka_dane.text()
        if file_path:
            self.load_file(file_path)

    def sprawdz_wpisy(self):
        miestac_roboczy = dodatki.data_miesiac_dzis()
        select_data = "SELECT * FROM `nieobecnosci_prod` WHERE miesiac = '%s';" % (miestac_roboczy)  # (miestac_roboczy)
        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        results = db.read_query(connection, select_data)
        connection.close()

        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        if results:
            for x in results:
                delete_data = "delete from nieobecnosci_prod where id = '%s' and Stanowisko is not Null and miesiac = '%s';" % (x[0],miestac_roboczy)
                print('Do skasowania:',delete_data)
                db.execute_query(connection, delete_data)
        else:
            print('--Brak wpisów jeszcze--')
        connection.close()

    def sprawdz_wpisy_obco(self):
        miestac_roboczy = dodatki.data_miesiac_dzis()
        select_data = "SELECT * FROM `nieobecnosci_prod` WHERE miesiac = '%s';" % (miestac_roboczy)  # (miestac_roboczy)
        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        results = db.read_query(connection, select_data)
        connection.close()

        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        if results:
            for x in results:
                delete_data = "delete from nieobecnosci_prod where id = '%s' and Stanowisko is Null and miesiac = '%s';" % (x[0],miestac_roboczy)
                print('Do skasowania:',delete_data)
                db.execute_query(connection, delete_data)
        else:
            print('--Brak wpisów jeszcze--')
        connection.close()

    def czytaj_dane_obco(self):
        if not self.folder_istnieje():
            return
        self.sprawdz_wpisy_obco()
        file_path = self.ui.ed_sciezka_dane.text()
        wb = openpyxl.load_workbook(os.path.join(file_path))
        sheet = wb.active
        used_columns = sheet.max_column
        if used_columns > 3:
            QMessageBox.critical(self, 'Error', 'Wybrałes zły plik. Ma za dużo wypełnionych kolumn!')
            return
        teraz = datetime.today()
        data_miesiac = str(dodatki.data_miesiac_dzis())
        print(data_miesiac)

        lista_wpisow = []

        #czytamy wszystkie kolumny i wiersze ze wskazanego pliku
        for row in sheet.iter_rows(min_row=2, min_col=0, max_col=3, values_only=True):
            #sprawdzamy czy wiersz nie jest pusty (zakładając że pusty wiersz ma wszystkie kolumny o wartosci None i kończy zestawienie)
            if any(cell is not None for cell in row):
                if row[1] == None:
                    break
                else:
                    dni_wolne = self.licz_dni_wolne(row[2])
                    lista_wpisow.append([row[0],row[1],row[2],dni_wolne,data_miesiac,teraz])
            else:
                break

        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)

        for row in lista_wpisow:
            insert_data = "INSERT INTO nieobecnosci_prod VALUES (NULL,'%s','%s',NULL,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,'%s','%s','%s','%s');" % (row[1],row[0],row[2],row[3],row[4],row[5])
            db.execute_query(connection, insert_data)

        self.load_data_from_database()

    def czytaj_dane(self):
        if not self.folder_istnieje():
            return
        self.sprawdz_wpisy()
        file_path = self.ui.ed_sciezka_dane.text()
        wb = openpyxl.load_workbook(os.path.join(file_path))
        sheet = wb.active
        used_columns = sheet.max_column
        if used_columns < 25:
            QMessageBox.critical(self, 'Error', 'Wybrałes zły plik. Ma za mało wypełnionych kolumn!')
            return
        teraz = datetime.today()
        data_miesiac = str(dodatki.data_miesiac_dzis())
        print(data_miesiac)

        lista_wpisow = []

        #czytamy wszystkie kolumny i wiersze ze wskazanego pliku
        for row in sheet.iter_rows(min_row=13, min_col=0, max_col=30, values_only=True):
            #sprawdzamy czy wiersz nie jest pusty (zakładając że pusty wiersz ma wszystkie kolumny o wartosci None i kończy zestawienie)
            if any(cell is not None for cell in row):
                if row[4] == None:
                    break
                else:
                    tekst = row[4].split('|')
                    nazwisko = tekst[0].strip()
                    nr_akt = tekst[1].strip()
                    #
                    lista_wpisow.append([nazwisko,nr_akt,row[6],row[8],row[9],row[10],row[11],row[12],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[26],row[27],row[29],data_miesiac,teraz])
                    #print(nazwisko,nr_akt,row[6],row[8],row[9],row[10],row[11],row[12],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[26],row[27],row[29],data_miesiac,teraz)
            else:
                break

        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        lista_wpisow_notNone = [tuple(0 if x is None else x for x in wiersz) for wiersz in lista_wpisow]

        for row in lista_wpisow_notNone:
            #print(dane)
            #row = [self.truncate_float(value) if isinstance(value, (int, float)) else value for value in dane]
            insert_data = "INSERT INTO nieobecnosci_prod VALUES (NULL,'%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s',0,'%s','%s','%s');" % (row[0],row[1],row[2],int(row[3]),int(row[4]),int(row[5]),int(row[6]),int(row[7]),int(row[8]),int(row[9]),int(row[10]),int(row[11]),int(row[12]),int(row[13]),int(row[14]),int(row[15]),int(row[16]),int(row[17]),int(row[18]),row[19],row[20])
            #print(insert_data)
            db.execute_query(connection, insert_data)

        self.load_data_from_database()

    def truncate_float(value):
        try:
            return math.floor(float(value))
        except ValueError:
            return value  # Jeśli to nie jest liczba, zwraca wartość oryginalną

    def folder_istnieje(self):
        folder = self.ui.ed_sciezka_dane.text().strip()
        if not folder:
            QMessageBox.critical(self, 'Error', 'Nie wybrano lokalizacji pliku')
            self.ui.ed_sciezka_dane.setFocus()
            return False

        if not os.path.exists(folder):
            QMessageBox.critical(self, 'Error', 'Folder nie istnieje. Sprawdź lokalizację pliku')
            self.ui.ed_sciezka_dane.setFocus()
            return False
        return True

    def licz_dni_wolne(self,dane):
        miestac_roboczy = dodatki.data_miesiac_dzis()
        data = datetime.strptime(miestac_roboczy, "%Y-%m-%d")
        miesiac = data.month
        rok = data.year
        _, dni_miesiaca = calendar.monthrange(rok, miesiac)

        select_data = "SELECT * FROM `dni_pracujace_w_roku` WHERE rok = '%s' and miesiac = '%s';" % (rok,miesiac)  # (miestac_roboczy)
        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        results = db.read_query(connection, select_data)

        dni_z_bazy = results[0][4]

        dni_wolne = dni_z_bazy - int(dane)
        if dni_wolne < 0:
            dni_wolne = 0
        return dni_wolne

    def show_filter_menu(self, col):
        values = set(
            self.ui.tab_dane.item(row, col).text()
            for row in range(self.ui.tab_dane.rowCount())
            if self.ui.tab_dane.item(row, col)
        )

        menu = QMenu(self)

        # Tworzenie obszaru przewijania
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_content = QFrame()
        layout = QVBoxLayout(scroll_content)

        checkboxes = {}
        max_width = 200  # Minimalna szerokość
        for value in sorted(values):
            checkbox = QCheckBox(value)
            checkbox.setChecked(
                col not in self.active_filters or value in self.active_filters[col]
            )
            layout.addWidget(checkbox)
            checkboxes[checkbox] = value

            # Dostosowanie szerokości na podstawie tekstu
            width = checkbox.fontMetrics().boundingRect(value).width() + 30
            max_width = min(max(max_width, width), 800)  # Maksymalna szerokość 800

        # Dodanie przycisków "Zaznacz wszystko" i "Wyczyść wszystko"
        button_select_all = QPushButton("Zaznacz wszystko")
        button_clear_all = QPushButton("Wyczyść wszystko")
        layout.addWidget(button_select_all)
        layout.addWidget(button_clear_all)

        button_select_all.clicked.connect(lambda: self.set_all_checkboxes(checkboxes, True))
        button_clear_all.clicked.connect(lambda: self.set_all_checkboxes(checkboxes, False))

        scroll_area.setWidget(scroll_content)
        scroll_area.setFixedHeight(200)  # Ograniczenie widocznych elementów do około 10 pozycji
        scroll_area.setMinimumWidth(max_width)
        scroll_area.setMaximumWidth(800)

        # Dodanie przewijanego obszaru do menu
        scroll_action = QWidgetAction(self)
        scroll_action.setDefaultWidget(scroll_area)
        menu.addAction(scroll_action)

        apply_action = menu.addAction("Zastosuj filtr")
        menu.addSeparator()
        clear_action = menu.addAction("Wyczyść filtr")

        header_pos = self.ui.tab_dane.mapToGlobal(self.ui.tab_dane.horizontalHeader().pos())
        section_pos = self.ui.tab_dane.horizontalHeader().sectionPosition(col)
        menu_pos = header_pos + QPoint(section_pos, self.ui.tab_dane.horizontalHeader().height())
        selected_action = menu.exec(menu_pos)

        if selected_action == apply_action:
            selected_values = [value for checkbox, value in checkboxes.items() if checkbox.isChecked()]
            self.active_filters[col] = selected_values
            self.apply_filters()
        elif selected_action == clear_action:
            if col in self.active_filters:
                del self.active_filters[col]
            self.apply_filters()

    def set_all_checkboxes(self, checkboxes, state):
        for checkbox in checkboxes.keys():
            checkbox.setChecked(state)

    def apply_filters(self):
        for row in range(self.ui.tab_dane.rowCount()):
            show_row = True
            for col, filter_values in self.active_filters.items():
                item = self.ui.tab_dane.item(row, col)
                if item and item.text() not in filter_values:
                    show_row = False
                    break
            self.ui.tab_dane.setRowHidden(row, not show_row)
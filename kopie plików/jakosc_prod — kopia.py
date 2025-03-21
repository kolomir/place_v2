from PyQt5.QtWidgets import QWidget, QMessageBox, QFileDialog, QTableWidgetItem,QHeaderView
import configparser
import openpyxl
#import sys
import os
from datetime import date, datetime
import calendar

from jakosc_prodDodaj import MainWindow_jakosc_prodDodaj
from _jakosc_prod_ui import Ui_Form
import db, dodatki


class MainWindow_jakosc(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        #- wczytanie pliku INI --------------------------------------------
        self.config = configparser.ConfigParser()
        self.config.read('config.ini')
        #- załadowanie zmiennych z pliku INI ------------------------------
        self.folder_bledy = self.config['sciezki']['folder_bledy']
        self.plik = self.config['sciezki']['plik_jakosc']
        #------------------------------------------------------------------
        # - domyślna ścieżka dla pliku -----------------
        #domyslny = f"{self.folder_bledy}"
        #self.ui.ed_sciezka_dane.setText(domyslny)
        # -----------------------------------------------

        self.ui.btn_przegladaj.clicked.connect(self.open_file_dialog)
        self.ui.btn_importuj.clicked.connect(self.czytaj_dane)
        self.ui.btn_dodaj.clicked.connect(self.otworz_okno_jakosc_prodDodaj)
        self.ui.btn_szablon.clicked.connect(self.szablon)
        self.wyszukaj_dane()

    def data_miesiac_dzis(self):
        data_dzis = date.today()
        prev_miesiac = data_dzis.month - 1 if data_dzis.month > 1 else 12
        prev_rok = data_dzis.year if data_dzis.month > 1 else data_dzis.year - 1
        data_miesiac = "%s-%s-%s" % (prev_rok,prev_miesiac,"1")
        print(data_miesiac)
        return data_miesiac

    def folder_istnieje(self):
        folder = self.ui.ed_sciezka_dane.text().strip()
        if not folder:
            QMessageBox.critical(self, 'Error', 'Nie wybrano lokalizacji pliku')
            self.ui.ed_sciezka_dane.setFocus()
            return False

        if not os.path.exists(folder):
            QMessageBox.critical(self, 'Error', 'Folder nie istnieje. Sprawdź lokalizację pliku')
            self.ui.ed_sciezka_dane.setFocus()
            return False
        return True

    def open_file_dialog(self):
        # Otwieranie dialogu wyboru pliku
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, "Wybierz plik Excel", "", "Pliki tekstowe (*.xlsx);;Wszystkie pliki (*)", options=options)
        if file_path:
            self.ui.ed_sciezka_dane.setText(file_path)  # Ustawienie ścieżki w polu tekstowym

    def load_from_path(self):
        # Wczytanie pliku z ręcznie wpisanej ścieżki
        file_path = self.ui.ed_sciezka_dane.text()
        if file_path:
            self.load_file(file_path)

    def czytaj_dane(self):
        if not self.folder_istnieje():
            return
        file_path = self.ui.ed_sciezka_dane.text()
        wb = openpyxl.load_workbook(os.path.join(file_path))
        sheet = wb.active
        teraz = datetime.today()
        data_miesiac = str(dodatki.data_miesiac_dzis())
        #print(data_miesiac)

        lista_wpisow = []

        i = 1
        #czytamy wszystkie kolumny i wiersze ze wskazanego pliku
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
            #sprawdzamy czy wiersz nie jest pusty (zakładając że pusty wiersz ma wszystkie kolumny o wartosci None i kończy zestawienie)
            if any(cell is not None for cell in row):
                grupa = row[0].strip()
                grupa_robocza = str(row[1]).strip()
                ppm = str(row[2]).strip()
                reklamacje = str(row[3]).strip()
                lista_wpisow.append([grupa, grupa_robocza, ppm, reklamacje, data_miesiac, teraz])
            else:
                break

        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)

        for row in lista_wpisow:
            insert_data = "INSERT INTO jakosc_prod VALUES (NULL,'%s','%s','%s','%s','%s','%s');" % (row[0], row[1], row[2], row[3], row[4], row[5])
            db.execute_query(connection, insert_data)

        self.wyszukaj_dane()

    def wyszukaj_dane(self):
        miestac_roboczy = dodatki.data_miesiac_dzis()
        print('miesiac',miestac_roboczy)
        select_data = "SELECT * FROM `jakosc_prod` WHERE miesiac = '%s';" % (miestac_roboczy) #(miestac_roboczy)
        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        results = db.read_query(connection, select_data)

        if not results:
            print('wynik ZLE')
            self.clear_table()
            self.naglowki_tabeli()
        else:
            print('wynik OK')
            self.naglowki_tabeli()
            self.pokaz_dane(results)
        connection.close()

    def clear_table(self):
        # Wyczyść zawartość tabeli
        self.ui.tab_dane.clearContents()
        self.ui.tab_dane.setRowCount(0)

    def naglowki_tabeli(self):
        self.ui.tab_dane.setColumnCount(6)  # Zmień na liczbę kolumn w twojej tabeli
        self.ui.tab_dane.setRowCount(0)  # Ustawienie liczby wierszy na 0
        self.ui.tab_dane.setHorizontalHeaderLabels(['Grupa', 'Grupa robocza', 'PPM', 'Reklamacje', 'Miesiąc', 'Data dodania'])

    def pokaz_dane(self, rows):
        # Column count
        if int(len(rows[0])) > 0:
            self.ui.tab_dane.setColumnCount(int(len(rows[0])) - 1)

        # Row count
        self.ui.tab_dane.setRowCount(int(len(rows)))

        wiersz = 0
        for wynik in rows:
            self.ui.tab_dane.setItem(wiersz, 0, QTableWidgetItem(str(wynik[1])))
            self.ui.tab_dane.setItem(wiersz, 1, QTableWidgetItem(str(wynik[2])))
            self.ui.tab_dane.setItem(wiersz, 2, QTableWidgetItem(str(wynik[3])))
            self.ui.tab_dane.setItem(wiersz, 3, QTableWidgetItem(str(wynik[4])))
            self.ui.tab_dane.setItem(wiersz, 4, QTableWidgetItem(str(wynik[5])))
            self.ui.tab_dane.setItem(wiersz, 5, QTableWidgetItem(str(wynik[6])))
            wiersz += 1

        self.ui.tab_dane.horizontalHeader().setStretchLastSection(True)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)

    def szablon(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        headers = ["Grupa","Grupa robocza","PPM","Reklamacje"]
        ws.append(headers)

        line1 = ["lider","3011","",""]
        line2 = ["lider","3012","",""]
        line3 = ["lider","3013 + 3015","",""]
        line4 = ["lider","3014","",""]
        line5 = ["lider","3016 + 3019","",""]
        line6 = ["lider","3018","",""]
        line7 = ["lider","3031+3032","",""]
        line8 = ["lider","Montaż","",""]
        line9 = ["lider","Zakuwanie","",""]
        line10 = ["lider","Maszyny","",""]
        line11 = ["instruktor","Czaplinek","",""]
        line12 = ["instruktor","Borne","",""]

        ws.append(line1)
        ws.append(line2)
        ws.append(line3)
        ws.append(line4)
        ws.append(line5)
        ws.append(line6)
        ws.append(line7)
        ws.append(line8)
        ws.append(line9)
        ws.append(line10)
        ws.append(line11)
        ws.append(line12)


        # Ustawianie szerokości kolumn
        ws.column_dimensions['A'].width = 15  # Grupa
        ws.column_dimensions['B'].width = 15  # Grupa robocza
        ws.column_dimensions['C'].width = 10  # PPM
        ws.column_dimensions['D'].width = 7  # Reklamacje

        domyslna_nazwa = 'jakosc.xlsx'
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            None,
            "Zapisz plik",
            domyslna_nazwa,  # Domyślna nazwa pliku
            "Pliki Excel (*.xlsx);;Wszystkie pliki (*)",
            options=options
        )

        # Sprawdzanie, czy użytkownik wybrał plik
        if file_path:
            wb.save(file_path)
            print(f"Plik zapisano: {file_path}")
        else:
            print("Zapis pliku anulowany")

    def otworz_okno_jakosc_prodDodaj(self):
        self.okno_jakosc_prodDodaj = MainWindow_jakosc_prodDodaj()
        self.okno_jakosc_prodDodaj.show()
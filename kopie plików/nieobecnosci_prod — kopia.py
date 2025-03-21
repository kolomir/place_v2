from PyQt5.QtWidgets import QWidget, QMessageBox, QFileDialog, QTableWidgetItem,QHeaderView
import configparser
import openpyxl
#import sys
import os
from datetime import date, datetime
import calendar

from plik_pomoc_nieobecnosci import MainWindow_pomoc_nieobecnosci
from _nieobecnosci_prod_ui import Ui_Form
import db, dodatki


class MainWindow_nieobecnosci(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        #- wczytanie pliku INI --------------------------------------------
        self.config = configparser.ConfigParser()
        self.config.read('config.ini')
        #- załadowanie zmiennych z pliku INI ------------------------------
        self.folder_bledy = self.config['sciezki']['folder_bledy']
        self.plik = self.config['sciezki']['plik_nieobecnosci']
        self.plik_obco = self.config['sciezki']['plik_obcokrajowcy']
        #------------------------------------------------------------------
        # - domyślna ścieżka dla pliku -----------------
        #domyslny = f"{self.folder_bledy}"
        #self.ui.ed_sciezka_dane.setText(domyslny)
        # -----------------------------------------------

        self.ui.btn_przegladaj.clicked.connect(self.open_file_dialog)
        self.ui.btn_importuj.clicked.connect(self.czytaj_dane)
        self.ui.btn_importuj_obco.clicked.connect(self.czytaj_dane_obco)
        self.ui.btn_szablon_p.clicked.connect(self.otworz_okno_plik_pomoc_nieobecnosci)
        self.ui.btn_szablon_o.clicked.connect(self.szablon_obco)
        self.wyszukaj_dane()


    def folder_istnieje(self):
        folder = self.ui.ed_sciezka_dane.text().strip()
        if not folder:
            QMessageBox.critical(self, 'Error', 'Nie wybrano lokalizacji pliku')
            self.ui.ed_sciezka_dane.setFocus()
            return False

        if not os.path.exists(folder):
            QMessageBox.critical(self, 'Error', 'Folder nie istnieje. Sprawdź lokalizację pliku')
            self.ui.ed_sciezka_dane.setFocus()
            return False
        return True

    def open_file_dialog(self):
        # Otwieranie dialogu wyboru pliku
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, "Wybierz plik Excel", "", "Pliki tekstowe (*.xlsx);;Wszystkie pliki (*)", options=options)
        if file_path:
            self.ui.ed_sciezka_dane.setText(file_path)  # Ustawienie ścieżki w polu tekstowym

    def load_from_path(self):
        # Wczytanie pliku z ręcznie wpisanej ścieżki
        file_path = self.ui.ed_sciezka_dane.text()
        if file_path:
            self.load_file(file_path)

    def licz_dni_wolne(self,dane):
        miestac_roboczy = dodatki.data_miesiac_dzis()
        data = datetime.strptime(miestac_roboczy, "%Y-%m-%d")
        miesiac = data.month
        rok = data.year
        _, dni_miesiaca = calendar.monthrange(rok, miesiac)

        select_data = "SELECT * FROM `dni_pracujace_w_roku` WHERE rok = '%s' and miesiac = '%s';" % (rok,miesiac)  # (miestac_roboczy)
        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        results = db.read_query(connection, select_data)

        dni_z_bazy = results[0][4]

        dni_wolne = dni_z_bazy - int(dane)
        if dni_wolne < 0:
            dni_wolne = 0
        return dni_wolne

    def sprawdz_wpisy(self):
        miestac_roboczy = dodatki.data_miesiac_dzis()
        select_data = "SELECT * FROM `nieobecnosci_prod` WHERE miesiac = '%s';" % (miestac_roboczy)  # (miestac_roboczy)
        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        results = db.read_query(connection, select_data)
        connection.close()

        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        if results:
            for x in results:
                delete_data = "delete from nieobecnosci_prod where id = '%s' and Stanowisko is not Null and miesiac = '%s';" % (x[0],miestac_roboczy)
                print('Do skasowania:',delete_data)
                db.execute_query(connection, delete_data)
        else:
            print('--Brak wpisów jeszcze--')
        connection.close()



    def sprawdz_wpisy_obco(self):
        miestac_roboczy = dodatki.data_miesiac_dzis()
        select_data = "SELECT * FROM `nieobecnosci_prod` WHERE miesiac = '%s';" % (miestac_roboczy)  # (miestac_roboczy)
        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        results = db.read_query(connection, select_data)
        connection.close()

        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        if results:
            for x in results:
                delete_data = "delete from nieobecnosci_prod where id = '%s' and Stanowisko is Null and miesiac = '%s';" % (x[0],miestac_roboczy)
                print('Do skasowania:',delete_data)
                db.execute_query(connection, delete_data)
        else:
            print('--Brak wpisów jeszcze--')
        connection.close()

    def czytaj_dane_obco(self):
        if not self.folder_istnieje():
            return
        self.sprawdz_wpisy_obco()
        file_path = self.ui.ed_sciezka_dane.text()
        wb = openpyxl.load_workbook(os.path.join(file_path))
        sheet = wb.active
        used_columns = sheet.max_column
        if used_columns > 3:
            QMessageBox.critical(self, 'Error', 'Wybrałes zły plik. Ma za dużo wypełnionych kolumn!')
            return
        teraz = datetime.today()
        data_miesiac = str(dodatki.data_miesiac_dzis())
        print(data_miesiac)

        lista_wpisow = []

        #czytamy wszystkie kolumny i wiersze ze wskazanego pliku
        for row in sheet.iter_rows(min_row=2, min_col=0, max_col=3, values_only=True):
            #sprawdzamy czy wiersz nie jest pusty (zakładając że pusty wiersz ma wszystkie kolumny o wartosci None i kończy zestawienie)
            if any(cell is not None for cell in row):
                if row[1] == None:
                    break
                else:
                    dni_wolne = self.licz_dni_wolne(row[2])
                    lista_wpisow.append([row[0],row[1],row[2],dni_wolne,data_miesiac,teraz])
            else:
                break

        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)

        for row in lista_wpisow:
            insert_data = "INSERT INTO nieobecnosci_prod VALUES (NULL,'%s','%s',NULL,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,'%s','%s','%s','%s');" % (row[1],row[0],row[2],row[3],row[4],row[5])
            db.execute_query(connection, insert_data)

        self.wyszukaj_dane()

    def czytaj_dane(self):
        if not self.folder_istnieje():
            return
        self.sprawdz_wpisy()
        file_path = self.ui.ed_sciezka_dane.text()
        wb = openpyxl.load_workbook(os.path.join(file_path))
        sheet = wb.active
        used_columns = sheet.max_column
        if used_columns < 25:
            QMessageBox.critical(self, 'Error', 'Wybrałes zły plik. Ma za mało wypełnionych kolumn!')
            return
        teraz = datetime.today()
        data_miesiac = str(dodatki.data_miesiac_dzis())
        print(data_miesiac)

        lista_wpisow = []

        #czytamy wszystkie kolumny i wiersze ze wskazanego pliku
        for row in sheet.iter_rows(min_row=13, min_col=0, max_col=30, values_only=True):
            #sprawdzamy czy wiersz nie jest pusty (zakładając że pusty wiersz ma wszystkie kolumny o wartosci None i kończy zestawienie)
            if any(cell is not None for cell in row):
                if row[4] == None:
                    break
                else:
                    tekst = row[4].split('|')
                    nazwisko = tekst[0].strip()
                    nr_akt = tekst[1].strip()
                    #
                    lista_wpisow.append([nazwisko,nr_akt,row[6],row[8],row[9],row[10],row[11],row[12],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[26],row[27],row[29],data_miesiac,teraz])
                    #print(nazwisko,nr_akt,row[6],row[8],row[9],row[10],row[11],row[12],row[17],row[18],row[19],row[20],row[21],row[22],row[23],row[24],row[26],row[27],row[29],data_miesiac,teraz)
            else:
                break

        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        lista_wpisow_notNone = [
            tuple(0 if x is None else x for x in wiersz) for wiersz in lista_wpisow
        ]

        for row in lista_wpisow_notNone:
            insert_data = "INSERT INTO nieobecnosci_prod VALUES (NULL,'%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s',0,'%s','%s','%s');" % (row[0],row[1],row[2],int(row[3]),int(row[4]),int(row[5]),int(row[6]),int(row[7]),int(row[8]),int(row[9]),int(row[10]),int(row[11]),int(row[12]),int(row[13]),int(row[14]),int(row[15]),int(row[16]),int(row[17]),int(row[18]),row[19],row[20])
            #print(insert_data)
            db.execute_query(connection, insert_data)

        self.wyszukaj_dane()

    def wyszukaj_dane(self):
        miestac_roboczy = dodatki.data_miesiac_dzis()
        print('miesiac',miestac_roboczy)
        select_data = "SELECT * FROM `nieobecnosci_prod` WHERE miesiac = '%s';" % (miestac_roboczy) #(miestac_roboczy)
        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        results = db.read_query(connection, select_data)

        if not results:
            print('wynik ZLE')
            self.clear_table()
            self.naglowki_tabeli()
        else:
            print('wynik OK')
            self.naglowki_tabeli()
            print(results)
            self.pokaz_dane(results)
        connection.close()

    def clear_table(self):
        # Wyczyść zawartość tabeli
        self.ui.tab_dane.clearContents()
        self.ui.tab_dane.setRowCount(0)

    def naglowki_tabeli(self):
        self.ui.tab_dane.setColumnCount(21)  # Zmień na liczbę kolumn w twojej tabeli
        self.ui.tab_dane.setRowCount(0)  # Ustawienie liczby wierszy na 0
        self.ui.tab_dane.setHorizontalHeaderLabels([
            'Nazwisko i imię',
            'Nr akt',
            'Stanowisko',
            'Urlop wypocz.',
            'Urlop bezpłatny',
            'Urlop szkoleniowy',
            'Urlop opieka (art. 188 kp) ',
            'Urlop okoliczność.',
            'Zwol. lek.',
            'Urlop macierz.',
            'Opieka ZUS',
            'Urlop wychow.',
            'Inne nieobecn.',
            'Usp.',
            'NN',
            'Rehab.',
            'Rodz.',
            'Krew',
            'Dni w pracy',
            'Razem',
            'Miesiac',
            'Data dodania'
        ])

    def pokaz_dane(self, rows):
        # Column count
        if int(len(rows[0])) > 0:
            self.ui.tab_dane.setColumnCount(int(len(rows[0])) - 1)

        # Row count
        self.ui.tab_dane.setRowCount(int(len(rows)))

        wiersz = 0
        for wynik in rows:
            self.ui.tab_dane.setItem(wiersz, 0, QTableWidgetItem(str(wynik[1])))
            self.ui.tab_dane.setItem(wiersz, 1, QTableWidgetItem(str(wynik[2])))
            self.ui.tab_dane.setItem(wiersz, 2, QTableWidgetItem(str(wynik[3])))
            self.ui.tab_dane.setItem(wiersz, 3, QTableWidgetItem(str(wynik[4])))
            self.ui.tab_dane.setItem(wiersz, 4, QTableWidgetItem(str(wynik[5])))
            self.ui.tab_dane.setItem(wiersz, 5, QTableWidgetItem(str(wynik[6])))
            self.ui.tab_dane.setItem(wiersz, 6, QTableWidgetItem(str(wynik[7])))
            self.ui.tab_dane.setItem(wiersz, 7, QTableWidgetItem(str(wynik[8])))
            self.ui.tab_dane.setItem(wiersz, 8, QTableWidgetItem(str(wynik[9])))
            self.ui.tab_dane.setItem(wiersz, 9, QTableWidgetItem(str(wynik[10])))
            self.ui.tab_dane.setItem(wiersz, 10, QTableWidgetItem(str(wynik[11])))
            self.ui.tab_dane.setItem(wiersz, 11, QTableWidgetItem(str(wynik[12])))
            self.ui.tab_dane.setItem(wiersz, 12, QTableWidgetItem(str(wynik[13])))
            self.ui.tab_dane.setItem(wiersz, 13, QTableWidgetItem(str(wynik[14])))
            self.ui.tab_dane.setItem(wiersz, 14, QTableWidgetItem(str(wynik[15])))
            self.ui.tab_dane.setItem(wiersz, 15, QTableWidgetItem(str(wynik[16])))
            self.ui.tab_dane.setItem(wiersz, 16, QTableWidgetItem(str(wynik[17])))
            self.ui.tab_dane.setItem(wiersz, 17, QTableWidgetItem(str(wynik[18])))
            self.ui.tab_dane.setItem(wiersz, 18, QTableWidgetItem(str(wynik[19])))
            self.ui.tab_dane.setItem(wiersz, 19, QTableWidgetItem(str(wynik[20])))
            self.ui.tab_dane.setItem(wiersz, 20, QTableWidgetItem(str(wynik[21])))
            self.ui.tab_dane.setItem(wiersz, 21, QTableWidgetItem(str(wynik[22])))
            wiersz += 1

        self.ui.tab_dane.horizontalHeader().setStretchLastSection(True)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(9, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(10, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(11, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(12, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(13, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(14, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(15, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(16, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(17, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(18, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(19, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(20, QHeaderView.ResizeToContents)
        self.ui.tab_dane.horizontalHeader().setSectionResizeMode(21, QHeaderView.ResizeToContents)

    def otworz_okno_plik_pomoc_nieobecnosci(self):
        self.okno_plik_pomoc_nieobecnosci = MainWindow_pomoc_nieobecnosci()
        self.okno_plik_pomoc_nieobecnosci.show()

    def szablon_obco(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        headers = ["nr_akt","Nazwisko i imie","dni w pracy"]
        ws.append(headers)

        # Ustawianie szerokości kolumn
        ws.column_dimensions['A'].width = 8  # Kolumna 'nr pracownika'
        ws.column_dimensions['B'].width = 30  # Kolumna 'IMIĘ i NAZWISKO'
        ws.column_dimensions['C'].width = 7  # Kolumna 'Suma z błędów'

        domyslna_nazwa = 'nieobecnosci_obcokrajowcow.xlsx'
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            None,
            "Zapisz plik",
            domyslna_nazwa,  # Domyślna nazwa pliku
            "Pliki Excel (*.xlsx);;Wszystkie pliki (*)",
            options=options
        )

        # Sprawdzanie, czy użytkownik wybrał plik
        if file_path:
            wb.save(file_path)
            print(f"Plik zapisano: {file_path}")
        else:
            print("Zapis pliku anulowany")

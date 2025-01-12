from PyQt5.QtWidgets import QWidget, QTableWidgetItem, QFileDialog, QMessageBox
from PyQt5.QtCore import Qt
import configparser
import openpyxl
#import sys
import os
from datetime import date, datetime

from _bledy_mag_ui import Ui_Form
import db, dodatki

# klasa która pozwala na poprawne sortowanie danych w kolumnach numerycznych.
# Normalnie dane układają się w sposób 1,10,11,2,23,245
# Po zastosowaniu klasy dane posortują się w sposób 1,2,10,11,23,245
class NumericTableWidgetItem(QTableWidgetItem):
    def __lt__(self, other):
        # Sprawdzamy, czy drugi element też jest instancją QTableWidgetItem
        if isinstance(other, QTableWidgetItem):
            try:
                # Porównujemy jako liczby
                return float(self.text()) < float(other.text())
            except ValueError:
                # W przypadku błędu porównujemy jako tekst
                return self.text() < other.text()
        return super().__lt__(other)

class MainWindow_bledy_mag(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        self.load_data_from_database()
        self.ui.tab_dane.itemChanged.connect(self.on_item_changed)
        self.ui.btn_przegladaj.clicked.connect(self.open_file_dialog)
        self.ui.btn_szablon.clicked.connect(self.szablon)
        self.ui.btn_importuj.clicked.connect(self.czytaj_dane)

    def open_file_dialog(self):
        # Otwieranie dialogu wyboru pliku
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, "Wybierz plik Excel", "", "Pliki tekstowe (*.xlsx);;Wszystkie pliki (*)", options=options)
        if file_path:
            self.ui.ed_sciezka_dane.setText(file_path)  # Ustawienie ścieżki w polu tekstowym

    def load_from_path(self):
        # Wczytanie pliku z ręcznie wpisanej ścieżki
        file_path = self.ui.ed_sciezka_dane.text()
        if file_path:
            self.load_file(file_path)

    def sprawdz_wpisy(self):
        miestac_roboczy = dodatki.data_miesiac_dzis()
        select_data = "SELECT * FROM `bledy_mag` WHERE miesiac = '%s';" % (miestac_roboczy)  # (miestac_roboczy)
        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        results = db.read_query(connection, select_data)
        connection.close()

        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
        if results:
            for x in results:
                delete_data = "delete from bledy_mag where id = '%s' and miesiac = '%s';" % (x[0],miestac_roboczy)
                #print('Do skasowania:',delete_data)
                db.execute_query(connection, delete_data)
        else:
            print('--Brak wpisów jeszcze--')
        connection.close()

    def czytaj_dane(self):
        if not self.folder_istnieje():
            return
        self.sprawdz_wpisy()
        file_path = self.ui.ed_sciezka_dane.text()
        wb = openpyxl.load_workbook(os.path.join(file_path))
        sheet = wb.active
        teraz = datetime.today()
        data_miesiac = str(dodatki.data_miesiac_dzis())
        print(data_miesiac)

        lista_wpisow = []

        # czytamy wszystkie kolumny i wiersze ze wskazanego pliku
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=5, values_only=True):
            # sprawdzamy czy wiersz nie jest pusty (zakładając że pusty wiersz ma wszystkie kolumny o wartosci None i kończy zestawienie)
            if any(cell is not None for cell in row):
                lista_wpisow.append([row[1], row[3], row[4], data_miesiac, teraz])
            else:
                break

        connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)

        for row in lista_wpisow:
            insert_data = "INSERT INTO bledy_mag VALUES (NULL,'%s','%s','%s','%s','%s');" % (
            row[0], row[1], row[2], row[3], row[4])
            db.execute_query(connection, insert_data)

        self.ui.tab_dane.blockSignals(True)
        self.load_data_from_database()
        self.ui.tab_dane.blockSignals(False)

    def load_data_from_database(self):
        try:
            miesiac_roboczy = dodatki.data_miesiac_dzis()
            select_data = "SELECT * FROM `bledy_mag` WHERE miesiac = '{0}';".format(miesiac_roboczy)
            connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
            results = db.read_query(connection, select_data)

            self.ui.tab_dane.setSortingEnabled(True)

            self.ui.tab_dane.setColumnCount(5)  # Zmień na liczbę kolumn w twojej tabeli
            self.ui.tab_dane.setRowCount(0)  # Ustawienie liczby wierszy na 0
            self.ui.tab_dane.setHorizontalHeaderLabels([
                'Nr akt',
                'Bledy zew.',
                'Bledy wew.',
                'Miesiac',
                'Data dodania'
            ])

            # Ustawianie liczby wierszy na podstawie danych z bazy
            self.ui.tab_dane.setRowCount(len(results))

            # Wypełnianie tabeli danymi
            for row_idx, row_data in enumerate(results):
                # Przechowujemy id każdego wiersza
                for col_idx, value in enumerate(row_data[1:]):  # Pomijamy id
                    item = NumericTableWidgetItem(str(value))              # Użycie klasy soryującej dane numeryczne

                    # Ustawianie wyrównania
                    if col_idx == 0 or col_idx == 1 or col_idx == 2 or col_idx == 3 or col_idx == 4:
                        item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                        # Kolumna 1: do lewej               (Qt.AlignLeft | Qt.AlignVCenter)
                        # Kolumna 2: wyśrodkowane           (Qt.AlignHCenter | Qt.AlignVCenter)
                        # Kolumna 3: do prawej              (Qt.AlignRight | Qt.AlignVCenter)

                    self.ui.tab_dane.setColumnWidth(0, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(1, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(2, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(3, 75)  # Stała szerokość: 150 pikseli
                    self.ui.tab_dane.setColumnWidth(4, 150)  # Stała szerokość: 150 pikseli


                    if col_idx == 0 or col_idx == 3 or col_idx == 4:  # Zablokowanie edycji dla kolumny "nazwa"
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Usuwamy flagę edytowalności
                    else:
                        item.setFlags(item.flags() | Qt.ItemIsEditable)  # Ustawienie komórek jako edytowalne
                    self.ui.tab_dane.setItem(row_idx, col_idx, item)



            # Przechowywanie id wierszy
            self.row_ids = [row_data[0] for row_data in results]
            #print(row_data[0] for row_data in results)

        except db.Error as e:
            print(f"Błąd przy pobieraniu danych z bazy danych: {e}")

    def update_database(self, record_id, col, new_value):
        """Funkcja do aktualizacji konkretnej komórki w bazie danych."""
        try:
            # Mapowanie indeksu kolumny na nazwę kolumny w bazie
            column_names = ["nr_akt", "bledy_zew", "bledy_wew"]
            column_name = column_names[col]
            print('col',col)
            print('column_name',column_name)

            # Aktualizacja w bazie danych
            sql_query = f"UPDATE bledy_mag SET {column_name} = %s WHERE id = %s"
            print('sql_query',sql_query)
            connection = db.create_db_connection(db.host_name, db.user_name, db.password, db.database_name)
            db.execute_query_virable(connection,sql_query,(new_value, record_id))
            #self.cursor.execute(sql_query, (new_value, record_id))
            #self.db_connection.commit()
            print(f"Zaktualizowano rekord o id {record_id}, {column_name} = {new_value}")

        except db.Error as e:
            print(f"Błąd zapisu do bazy danych: {e}")

    def on_item_changed(self, item):
        """Funkcja wywoływana przy każdej zmianie komórki."""
        row = item.row()
        col = item.column()
        # print(f"DEBUG: Wybrane kolumny: {col}")
        new_value = item.text()

        # Pobranie id rekordu dla zmienionego wiersza
        record_id = self.row_ids[row]
        # print('record_id:',record_id)

        # Zapis zmienionych danych do bazy
        self.update_database(record_id, col, new_value)

    def szablon(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        headers = ["lp","nr prac.","imię i nazwisko","Błędy zewnętrzne","Błędy wewnętrzne"]
        ws.append(headers)

        # Ustawianie szerokości kolumn
        ws.column_dimensions['A'].width = 5  # Kolumna 'lp'
        ws.column_dimensions['B'].width = 8  # Kolumna 'nr pracownika'
        ws.column_dimensions['C'].width = 30  # Kolumna 'IMIĘ i NAZWISKO'
        ws.column_dimensions['D'].width = 5  # Kolumna 'Błędy zewnętrzne'
        ws.column_dimensions['E'].width = 5  # Kolumna 'Błędy wewnętrzne'

        domyslna_nazwa = 'bledy_mag.xlsx'
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            None,
            "Zapisz plik",
            domyslna_nazwa,  # Domyślna nazwa pliku
            "Pliki Excel (*.xlsx);;Wszystkie pliki (*)",
            options=options
        )

        # Sprawdzanie, czy użytkownik wybrał plik
        if file_path:
            wb.save(file_path)
            print(f"Plik zapisano: {file_path}")
        else:
            print("Zapis pliku anulowany")

    def folder_istnieje(self):
        folder = self.ui.ed_sciezka_dane.text().strip()
        if not folder:
            QMessageBox.critical(self, 'Error', 'Nie wybrano lokalizacji pliku')
            self.ui.ed_sciezka_dane.setFocus()
            return False

        if not os.path.exists(folder):
            QMessageBox.critical(self, 'Error', 'Folder nie istnieje. Sprawdź lokalizację pliku')
            self.ui.ed_sciezka_dane.setFocus()
            return False
        return True